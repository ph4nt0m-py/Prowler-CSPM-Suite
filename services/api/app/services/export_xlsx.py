"""Export findings for a scan to XLSX (openpyxl)."""

from __future__ import annotations

import io
import uuid
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from app.models.finding import Finding
from app.models.scan import Scan
from app.models.triage import FindingTriage


def build_findings_xlsx(db: Session, scan_id: uuid.UUID) -> bytes:
    scan = db.get(Scan, scan_id)
    if not scan:
        raise ValueError("Scan not found")

    triage_map: dict[str, Any] = {
        t.fingerprint: t.state.value
        for t in db.query(FindingTriage).filter(FindingTriage.client_id == scan.client_id).all()
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Findings"
    headers = [
        "Severity",
        "Status",
        "Triage",
        "Description",
        "Resource",
        "Region",
        "Service",
        "Check ID",
        "Compliance",
        "Fingerprint",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    findings = db.query(Finding).filter(Finding.scan_id == scan_id).order_by(Finding.severity.desc()).all()
    row = 2
    for f in findings:
        ws.cell(row=row, column=1, value=f.severity.value)
        ws.cell(row=row, column=2, value=f.status.value)
        ws.cell(row=row, column=3, value=triage_map.get(f.fingerprint, ""))
        ws.cell(row=row, column=4, value=f.description or "")
        ws.cell(row=row, column=5, value=f.resource_id)
        ws.cell(row=row, column=6, value=f.region)
        ws.cell(row=row, column=7, value=f.service)
        ws.cell(row=row, column=8, value=f.check_id)
        ws.cell(row=row, column=9, value=f.compliance_framework or "")
        ws.cell(row=row, column=10, value=f.fingerprint)
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
